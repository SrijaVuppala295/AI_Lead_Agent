"""
excel_writer.py — Single Sheet Excel Writer
=============================================
ONE sheet: "Leads"
One row per lead. Every email has its own Subject + Body + Status columns.

COLUMN GROUPS:
  Lead Info (navy)      : Company, Contact, Title, LinkedIn, Website, Email, Source, Signal, Score, Date
  Original (blue)       : Subject | Body | Status
  Followup 1 (purple)   : Subject | Body | Status
  Followup 2 (purple)   : Subject | Body | Status
  Followup 3 (purple)   : Subject | Body | Status
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.database import get_all_leads, get_emails_by_status
from datetime import datetime
import re, os

OUTPUT_DIR  = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "leads.xlsx")

HEADER_BG = "1F3864"   # navy   — lead cols
ORIG_HDR  = "1565C0"   # blue   — original email
FU_HDR    = "6A1B9A"   # purple — followups
HOT_BG    = "FFE0E0"
WARM_BG   = "FFFDE7"
ALT_BG    = "F5F5F5"
EMAIL_BG  = "E8F5E9"   # green  — has email
ORIG_BG   = "E3F2FD"   # light blue  — original cells
FU_BG     = "F3E5F5"   # light purple — followup cells
TBD_FG    = "AAAAAA"

# fmt: off
COLUMNS = [
    # key,                     header,                  width,  group
    ("company_name",           "Company Name",           22,    "lead"),
    ("contact_name",           "Contact Name",           18,    "lead"),
    ("contact_title",          "Title",                  22,    "lead"),
    ("linkedin_url",           "LinkedIn URL",           32,    "lead"),
    ("company_website",        "Company Website",        25,    "lead"),
    ("contact_email",          "Contact Email",          28,    "lead"),
    ("source_name",            "Signal Source",          16,    "lead"),
    ("signal_title",           "Signal Summary",         40,    "lead"),
    ("urgency_score",          "Intent Score",           13,    "lead"),
    ("date_found",             "Date Found",             18,    "lead"),
    # Original
    ("orig_subject",           "Original Subject",       30,    "orig"),
    ("orig_body",              "Original Email Body",    55,    "orig"),
    ("orig_status",            "Original Status",        18,    "orig"),
    # Followup 1
    ("fu1_subject",            "Followup 1 Subject",     30,    "fu"),
    ("fu1_body",               "Followup 1 Body",        55,    "fu"),
    ("fu1_status",             "Followup 1 Status",      18,    "fu"),
    # Followup 2
    ("fu2_subject",            "Followup 2 Subject",     30,    "fu"),
    ("fu2_body",               "Followup 2 Body",        55,    "fu"),
    ("fu2_status",             "Followup 2 Status",      18,    "fu"),
    # Followup 3
    ("fu3_subject",            "Followup 3 Subject",     30,    "fu"),
    ("fu3_body",               "Followup 3 Body",        55,    "fu"),
    ("fu3_status",             "Followup 3 Status",      18,    "fu"),
]
# fmt: on

COL_INDEX = {col[0]: i for i, col in enumerate(COLUMNS, 1)}
HDR_BG    = {"lead": HEADER_BG, "orig": ORIG_HDR, "fu": FU_HDR}


def _clean_email(email):
    if not email or "@" not in email:
        return email
    m = re.match(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,6}", email)
    return m.group(0) if m else email


def _email_status(em):
    """Returns human-readable status for one email row."""
    if em is None:
        return "not generated"
    s = em[8] or "generated"
    if s == "replied":   return "replied ✅"
    if s == "sent":      return "sent ✉️"
    if s == "generated": return "generated ✅"
    return s


def _thin():
    t = Side(style="thin", color="DDDDDD")
    return Border(bottom=t, right=t)


def _hdr(cell, value, group):
    bg = HDR_BG.get(group, HEADER_BG)
    cell.value     = value
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(bottom=Side(style="medium", color="FFFFFF"),
                            right=Side(style="medium", color="FFFFFF"))


def _cell(cell, value, key, row_bg):
    # Email body/status cells get their own bg
    if key in ("orig_subject","orig_body","orig_status"):
        bg = ORIG_BG
    elif key.startswith("fu") and key.endswith(("subject","body","status")):
        bg = FU_BG
    else:
        bg = row_bg

    cell.value     = value
    if bg:
        cell.fill  = PatternFill("solid", fgColor=bg)
    cell.border    = _thin()
    cell.alignment = Alignment(
        vertical="top", wrap_text=True,
        horizontal="center" if key == "urgency_score" else "left"
    )

    is_tbd = str(value or "") in ("TBD","Not found","not generated","")
    if is_tbd:
        cell.font = Font(name="Arial", size=9, italic=True, color=TBD_FG)
    elif key == "company_name":
        cell.font = Font(name="Arial", size=10, bold=True)
    elif key == "urgency_score":
        clr = "C00000" if (value or 0) >= 7 else ("7B6000" if (value or 0) >= 5 else "000000")
        cell.font = Font(name="Arial", size=10, bold=True, color=clr)
    elif key == "contact_email" and value and value not in ("TBD","Not found",""):
        cell.font = Font(name="Arial", size=9, color="1565C0")
    elif key.endswith("_status") and value and "✅" in str(value):
        cell.font = Font(name="Arial", size=9, bold=True, color="2E7D32")
    elif key.endswith("_status") and value and "✉️" in str(value):
        cell.font = Font(name="Arial", size=9, bold=True, color="1565C0")
    elif key.endswith("_body"):
        cell.font = Font(name="Arial", size=9)
    else:
        cell.font = Font(name="Arial", size=10)


# ── Real-time helpers (just rebuild each time) ─────────────────────────────
def _ensure_file():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def write_lead_row(lead_data):        build_leads_xlsx()
def write_contact_to_row(cn, c):      build_leads_xlsx()
def write_email_to_row(cn, e, conf):  build_leads_xlsx()
def write_email_sent_to_row(cn, et):  build_leads_xlsx()
def write_reply_to_row(cn):           build_leads_xlsx()


# ═══════════════════════════════════════════════════════════════════════════
# FULL REBUILD
# ═══════════════════════════════════════════════════════════════════════════

def build_leads_xlsx():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    leads  = get_all_leads()
    emails = get_emails_by_status()

    # Index: lead_id → {email_type: row}
    email_idx = {}
    for em in emails:
        lid, etype = em[1], em[7] or ""
        email_idx.setdefault(lid, {})[etype] = em

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header
    for i, (key, header, width, group) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i)
        _hdr(cell, header, group)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, lead in enumerate(leads, 2):
        lead_id         = lead[0]
        company_name    = lead[1]  or ""
        signal_title    = lead[2]  or ""
        urgency_score   = lead[4]  or 0
        source_name     = lead[7]  or ""
        date_found      = str(lead[8] or "")
        contact_name    = lead[9]  if len(lead) > 9  else "TBD"
        contact_title   = lead[10] if len(lead) > 10 else "TBD"
        company_website = lead[11] if len(lead) > 11 else "TBD"
        linkedin_url    = lead[12] if len(lead) > 12 else "TBD"
        contact_email   = _clean_email(lead[13]) if len(lead) > 13 and lead[13] else ""

        le   = email_idx.get(lead_id, {})
        orig = le.get("original")
        fu1  = le.get("followup_1")
        fu2  = le.get("followup_2")
        fu3  = le.get("followup_3")

        has_emails = bool(le)
        row_bg = (EMAIL_BG if has_emails
                  else HOT_BG  if urgency_score >= 7
                  else WARM_BG if urgency_score >= 5
                  else ALT_BG  if row_idx % 2 == 0
                  else None)

        row_data = {
            "company_name":    company_name,
            "contact_name":    contact_name    or "TBD",
            "contact_title":   contact_title   or "TBD",
            "linkedin_url":    linkedin_url    or "TBD",
            "company_website": company_website or "TBD",
            "contact_email":   contact_email   or "TBD",
            "source_name":     source_name,
            "signal_title":    signal_title,
            "urgency_score":   urgency_score,
            "date_found":      date_found,
            # Original
            "orig_subject": orig[5] if orig else "",
            "orig_body":    orig[6] if orig else "",
            "orig_status":  _email_status(orig),
            # Followup 1
            "fu1_subject":  fu1[5] if fu1 else "",
            "fu1_body":     fu1[6] if fu1 else "",
            "fu1_status":   _email_status(fu1),
            # Followup 2
            "fu2_subject":  fu2[5] if fu2 else "",
            "fu2_body":     fu2[6] if fu2 else "",
            "fu2_status":   _email_status(fu2),
            # Followup 3
            "fu3_subject":  fu3[5] if fu3 else "",
            "fu3_body":     fu3[6] if fu3 else "",
            "fu3_status":   _email_status(fu3),
        }

        for i, (key, header, width, group) in enumerate(COLUMNS, 1):
            _cell(ws.cell(row=row_idx, column=i), row_data[key], key, row_bg)

        ws.row_dimensions[row_idx].height = 120 if has_emails else 50

    if not leads:
        wb.save(OUTPUT_FILE)
        return None

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE